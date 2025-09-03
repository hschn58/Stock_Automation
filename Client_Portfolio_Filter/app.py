import re
import io
import base64
import pandas as pd
from flask import Flask, request, render_template_string, send_file, session, redirect, url_for
import matplotlib

matplotlib.use("Agg")  # Use a non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import os
import time

from io import BytesIO
import openpyxl
import defusedxml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import json

import sys
import threading
import webbrowser

from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox
from PyQt5.QtWebEngineWidgets import QWebEngineView, QWebEngineProfile
from PyQt5.QtCore import QUrl

from collections import defaultdict

from functools import lru_cache  # NEW
from flask import make_response

from datetime import datetime  # 01 NEW# NEW (optional for headers)

app = Flask(__name__)

app.secret_key = os.environ.get("APP_SECRET_KEY", os.urandom(32))

from flask_cors import CORS

CORS(app)


# --- Stable color mapping across the session ------------------------------------------  # 01
# Customize these to hard-pin colors you care about (keys are case-insensitive).         # 02
PREFERRED_CLASS_COLORS = {  # 03
    # e.g., 'cash and equivalents': '#8c564b',                                           # 04
}  # 05
PREFERRED_SEGMENT_COLORS = {  # 06
    "technology": "#1f77b4",  # blue                                                    # 07
    "utilities": "#9467bd",  # purple                                                  # 08
}  # 10

# Per-session registries filled as new labels appear                                      # 11
SESSION_CLASS_COLORS = {}  # normalized label -> hex                                  # 12
SESSION_SEGMENT_COLORS = {}  # normalized label -> hex                                  # 13


def _norm_label(s):  # 14
    s = str(s).strip().lower()  # 15
    s = re.sub(r"[^a-z0-9 ]+", "", s)  # letters/digits/spaces only                     # 16
    s = re.sub(r"\s+", " ", s)  # 17
    return s  # 18


# A stable palette big enough for many categories                                         # 19
def _palette_20():  # 20
    try:  # 21
        tab20 = plt.get_cmap("tab20")  # 22
        return [mcolors.to_hex(tab20(i)) for i in range(tab20.N)]  # 23
    except Exception:  # 24
        return list(mcolors.TABLEAU_COLORS.values())  # 25


_BASE_PALETTE = _palette_20()  # 26


def _assign_colors(labels, *, kind):  # 27
    """Return a color list aligned with `labels`, using per-session stable mapping."""  # 28
    if kind == "class":  # 29
        pref, reg = PREFERRED_CLASS_COLORS, SESSION_CLASS_COLORS  # 30
    else:  # 31
        pref, reg = PREFERRED_SEGMENT_COLORS, SESSION_SEGMENT_COLORS  # 32
    used = set(reg.values()) | {
        pref.get(k) for k in map(_norm_label, pref.keys()) if pref.get(k)
    }  # 33
    colors = []  # 34
    for lbl in labels:  # 35
        key = _norm_label(lbl)  # 36
        # 1) preferred                                                                     # 37
        if key in {_norm_label(k): None for k in pref.keys()}:  # 38
            # find the exact color from pref by original key lookup                       # 39
            for k, v in pref.items():  # 40
                if _norm_label(k) == key:  # 41
                    reg[key] = v  # 42
                    used.add(v)  # 43
                    colors.append(v)  # 44
                    break  # 45
            continue  # 46
        # 2) already assigned                                                              # 47
        if key in reg:  # 48
            colors.append(reg[key])  # 49
            continue  # 50
        # 3) pick next unused from palette                                                 # 51
        col = next((c for c in _BASE_PALETTE if c not in used), None)  # 52
        if col is None:
            # fallback: deterministic hash pick                                            # 53
            col = _BASE_PALETTE[abs(hash(key)) % len(_BASE_PALETTE)]  # 54
        reg[key] = col  # 55
        used.add(col)  # 56
        colors.append(col)  # 57
    return colors  # 58


def prime_color_maps(all_classes, all_segments):  # 59
    """Optional: pre-assign colors so first chart already uses stable colors."""  # 60
    if all_classes:
        _assign_colors(sorted(all_classes, key=lambda x: str(x).lower()), kind="class")  # 61
    if all_segments:
        _assign_colors(sorted(all_segments, key=lambda x: str(x).lower()), kind="segment")  # 62


# --------------------------------------------------------------------------------------  # 63

###########################################################################################

# Safe defaults so the first GET never fails even before data loads
COMMON_STOCK = "COMMON STOCK"
CASH_CLASS = "CASH AND EQUIVALENTS"

# --- Two-CSV configuration -----------------------------------------------------------  # 01
CSV_ONE_FILENAME = os.environ.get("CSV_ONE_FILENAME", "Client_Data.csv")  # 02
CSV_TWO_FILENAME = os.environ.get("CSV_TWO_FILENAME", "ShortName_Map.csv")  # 03
CSV1_PATH, CSV2_PATH = None, None  # 04
# -------------------------------------------------------------------------------------  # 05


def read_csv_any_encoding(path):  # 06
    """Read a CSV trying common encodings; raise if all fail."""  # 07
    for enc in (None, "utf-8", "utf-8-sig", "cp1252", "latin1"):  # 08
        try:  # 09
            return pd.read_csv(path) if enc is None else pd.read_csv(path, encoding=enc)  # 10
        except UnicodeDecodeError:  # 11
            continue  # 12
    # last attempt throws
    return pd.read_csv(path)  # 13


def _norm_token(s):  # 14
    return re.sub(r"[^a-z0-9]+", "", str(s).lower())  # 15


def find_col(columns, *candidates, required=False):  # 16
    """Try search_pattern first, then tolerant substring matching."""  # 17
    cols = list(columns)  # 18
    # 1) try exact-ish search_pattern                                                   # 19
    for cand in candidates:  # 20
        col = search_pattern(cols, cand)  # 21
        if col:
            return col  # 22
    # 2) tolerant substring on normalized tokens                                        # 23
    norm_cols = {_norm_token(c): c for c in cols}  # 24
    for cand in candidates:  # 25
        nc = _norm_token(cand)  # 26
        for c in cols:  # 27
            if nc in _norm_token(c):  # 28
                return c  # 29
    if required:  # 30
        raise KeyError(f"Could not find any of {candidates} in columns: {cols}")  # 31
    return None  # 32


def normalize_account_series(s):  # 33
    """Keep leading zeros; strip surrounding whitespace. Customize if needed."""  # 34
    return s.astype(str).str.strip()  # 35


#############################################################################################################


def search_pattern(strings, pattern):
    """
    Return the first matching string (case-insensitive, ignoring non-letters),
    or None if nothing matches.
    """
    normalized_pattern = re.sub(r"\s+", " ", pattern.strip()).lower()
    pattern_regex = re.compile(re.escape(normalized_pattern), re.IGNORECASE)

    matches = []
    for string in strings:
        normalized_string = re.sub(r"[^a-zA-Z ]", "", str(string)).lower()
        if pattern_regex.search(normalized_string):
            matches.append(string)

    return matches[0] if matches else None


# --- Columns ---

DATA_FILENAME = "Client_Data.csv"
CSV_PATH = None  # global that gets set once the file is confirmed
data = None  # maybe store a global DataFrame
# --- Build cache lists ---

# ------------- HTML TEMPLATE -------------
from template import template


def filter_by_sector(df, sector, operator_, target_str):
    """
    Keep only rows belonging to (Portfolio, Short Name) pairs whose share of the selected
    sector meets the operator/target. Percentage is sector_value / total_value for the pair.
    """
    if not sector:
        return df

    try:
        target_percent = float(target_str)
    except ValueError:
        return pd.DataFrame(columns=df.columns)

        # inside filter_by_sector(...) after target_percent = float(target_str)
    if target_percent < 0:  # 01
        return pd.DataFrame(columns=df.columns)  # 02

    # Totals by (Portfolio, Short Name)
    totals = (
        df.groupby([COL_PORTFOLIO, COL_SHORTNAME], as_index=False)[COL_VALUE]
        .sum()
        .rename(columns={COL_VALUE: "total_value"})
    )

    # Sector sum by (Portfolio, Short Name)
    sec_sum = (
        df[df[COL_SEGMENT] == sector]
        .groupby([COL_PORTFOLIO, COL_SHORTNAME], as_index=False)[COL_VALUE]
        .sum()
        .rename(columns={COL_VALUE: "sector_value"})
    )

    merged = totals.merge(sec_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how="left")
    merged["sector_value"] = merged["sector_value"].fillna(0.0)
    merged["pct"] = (merged["sector_value"] / merged["total_value"]) * 100.0

    if operator_ == "lt":
        keep = merged[merged["pct"] < target_percent]
    elif operator_ == "gt":
        keep = merged[merged["pct"] > target_percent]
    elif operator_ == "eq":
        keep = merged[merged["pct"].sub(target_percent).abs() <= 3]
    else:
        return pd.DataFrame(columns=df.columns)

    keep_pairs = keep[[COL_PORTFOLIO, COL_SHORTNAME]].drop_duplicates()
    out = df.merge(
        keep_pairs.assign(__keep=1), on=[COL_PORTFOLIO, COL_SHORTNAME], how="inner"
    ).drop(columns="__keep")
    return out


def filter_single(df, class_filter, operator_, target_str):
    """Filter at (Portfolio, Short Name) level, keeping only rows for passing pairs."""
    if not class_filter:
        return df  # No filter => pass-through

    try:
        target_percent = float(target_str)
    except ValueError:
        return pd.DataFrame(columns=df.columns)

        # inside filter_single(...) after target_percent = float(target_str)
    if target_percent < 0:  # 01
        return pd.DataFrame(columns=df.columns)  # 02

    # Totals by (Portfolio, Short Name)
    totals = (
        df.groupby([COL_PORTFOLIO, COL_SHORTNAME], as_index=False)[COL_VALUE]
        .sum()
        .rename(columns={COL_VALUE: "total_value"})
    )

    # Class sum by (Portfolio, Short Name)
    cls_sum = (
        df[df[COL_CLASS] == class_filter]
        .groupby([COL_PORTFOLIO, COL_SHORTNAME], as_index=False)[COL_VALUE]
        .sum()
        .rename(columns={COL_VALUE: "class_value"})
    )

    merged = totals.merge(cls_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how="left")
    merged["class_value"] = merged["class_value"].fillna(0.0)
    merged["pct"] = (merged["class_value"] / merged["total_value"]) * 100.0

    if operator_ == "lt":
        keep = merged[merged["pct"] < target_percent]
    elif operator_ == "gt":
        keep = merged[merged["pct"] > target_percent]
    elif operator_ == "eq":
        keep = merged[merged["pct"].sub(target_percent).abs() <= 3]
    else:
        return pd.DataFrame(columns=df.columns)

    keep_pairs = keep[[COL_PORTFOLIO, COL_SHORTNAME]].drop_duplicates()
    out = df.merge(
        keep_pairs.assign(__keep=1), on=[COL_PORTFOLIO, COL_SHORTNAME], how="inner"
    ).drop(columns="__keep")
    return out


def apply_both_filters(class1, op1, pct1, class2, op2, pct2, use_second, df):
    """Apply up to 2 filters in an AND fashion on df. If use_second=True, apply second also."""
    df1 = filter_single(df, class1, op1, pct1)

    if not use_second:

        total_by_portfolio = df1.groupby(COL_PORTFOLIO)[COL_VALUE].sum().rename("Portfolio Value")
        df1 = df1.merge(total_by_portfolio, on=COL_PORTFOLIO, how="left")

        return df1
    # else, filter the result again
    df2 = filter_single(df1, class2, op2, pct2)

    # Calculate portfolio value
    total_by_portfolio = df2.groupby(COL_PORTFOLIO)[COL_VALUE].sum().rename("Portfolio Value")
    df2 = df2.merge(total_by_portfolio, on=COL_PORTFOLIO, how="left")

    return df2


def convert_to_percent(data_list):
    total = sum(data_list)
    if total == 0:
        return [0 for _ in data_list]
    return [round((val / total) * 100.0, 2) for val in data_list]


#################################################################################
def file_created_str(path):  # 01 NEW
    """Return a human-readable local 'created' timestamp for a file.               # 02 NEW
    On macOS/Windows uses birthtime; elsewhere falls back to modification time."""  # 03 NEW
    try:  # 04 NEW
        st = os.stat(path)  # 05 NEW
        if hasattr(st, "st_birthtime") and st.st_birthtime:  # 06 NEW
            ts = st.st_birthtime  # macOS (and some BSDs)                          # 07 NEW
        elif sys.platform.startswith("win"):  # 08 NEW
            ts = st.st_ctime  # Windows creation time                           # 09 NEW
        else:
            ts = st.st_mtime  # Fallback: last modified                         # 10 NEW
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d")  # 11 NEW
    except Exception:  # 12 NEW
        return "N/A"  # 13 NEW


#################################################################################


# 36
def generate_pie_chart(portfolio, shortName=None, threshold=3):  # 01
    """Generate a pie chart of Class breakdown for a portfolio (and optional shortName)."""  # 02
    df_filtered = data[data[COL_PORTFOLIO] == portfolio]  # 03
    if shortName:  # 04
        df_filtered = df_filtered[df_filtered[COL_SHORTNAME] == shortName]  # 05
    if df_filtered.empty:  # 06
        return None  # 07
    class_values = (
        df_filtered.groupby(COL_CLASS)[COL_VALUE].sum().sort_values(ascending=False)  # 08
    )  # 09
    labels = list(class_values.index)  # 10
    colors = _assign_colors(labels, kind="class")  # 11
    # 12
    fig, ax = plt.subplots(figsize=(5.4, 5.4))  # 13
    ax.set_aspect("equal")  # 14
    wedges, _texts, _autotexts = ax.pie(  # 15
        class_values.values,  # 16
        labels=None,  # 17
        autopct=lambda p: f"{p:.1f}%" if p >= threshold else "",  # 18
        startangle=90,  # 19
        radius=1.0,  # 20
        colors=colors,  # 21
    )  # 22
    title_name = shortName if shortName else str(portfolio)  # 23
    ax.set_title(f"Asset Breakdown — {title_name}", pad=12)  # 24
    legend_labels = [  # 25
        f"{lbl} ({pct}%)" if pct <= threshold else f"{lbl}"  # 26
        for lbl, pct in zip(labels, convert_to_percent(class_values.values))  # 27
    ]  # 28
    fig.subplots_adjust(left=0.06, right=0.94, top=0.86, bottom=0.25)  # 29
    fig.legend(
        wedges,
        legend_labels,
        title="Class",
        loc="lower center",  # 30
        bbox_to_anchor=(0.5, 0.02),
        ncol=2,
        frameon=True,
    )  # 31
    buf = io.BytesIO()  # 32
    fig.savefig(buf, format="png", dpi=144)  # 33
    buf.seek(0)
    plt.close(fig)  # 34
    return base64.b64encode(buf.getvalue()).decode("utf-8")  # 35


def generate_sector_pie_chart(portfolio, short_name=None, threshold=2.5):  # 01
    """Among COMMON STOCK holdings, break down by Segment (Sector) in a separate pie."""  # 02
    df_filtered = data[(data[COL_PORTFOLIO] == portfolio) & (data[COL_CLASS] == COMMON_STOCK)]  # 03
    if short_name:  # 04
        df_filtered = df_filtered[df_filtered[COL_SHORTNAME] == short_name]  # 05
    if df_filtered.empty:  # 06
        return None  # 07
    sector_values = (
        df_filtered.groupby(COL_SEGMENT)[COL_VALUE].sum().sort_values(ascending=False)  # 08
    )  # 09
    labels = list(sector_values.index)  # 10
    colors = _assign_colors(labels, kind="segment")  # 11
    # 12
    fig, ax = plt.subplots(figsize=(5.4, 5.4))  # 13
    ax.set_aspect("equal")  # 14
    wedges, _txt, _autotxt = ax.pie(  # 15
        sector_values.values,  # 16
        labels=None,  # 17
        autopct=lambda p: f"{p:.1f}%" if p >= threshold else "",  # 18
        startangle=90,  # 19
        radius=1.0,  # 20
        colors=colors,  # 21
    )  # 22
    title_name = short_name if short_name else str(portfolio)  # 23
    ax.set_title(f"Common Stock by Sector — {title_name}", pad=12)  # 24
    legend_labels = [  # 25
        f"{lbl} ({pct}%)" if pct <= threshold else f"{lbl}"  # 26
        for lbl, pct in zip(labels, convert_to_percent(sector_values.values))  # 27
    ]  # 28
    fig.subplots_adjust(left=0.06, right=0.94, top=0.86, bottom=0.25)  # 29
    fig.legend(
        wedges,
        legend_labels,
        title="Segment",
        loc="lower center",  # 30
        bbox_to_anchor=(0.5, 0.02),
        ncol=2,
        frameon=True,
    )  # 31
    buf = io.BytesIO()  # 32
    fig.savefig(buf, format="png", dpi=144)  # 33
    buf.seek(0)
    plt.close(fig)  # 34
    return base64.b64encode(buf.getvalue()).decode("utf-8")  # 35


@lru_cache(maxsize=256)
def _pie_bytes(portfolio, short_name, threshold=3):  # 01
    b64 = generate_pie_chart(portfolio, short_name, threshold)  # 02
    return base64.b64decode(b64) if b64 else None  # 03


@lru_cache(maxsize=256)
def _sector_pie_bytes(portfolio, short_name, threshold=2.5):  # 04
    b64 = generate_sector_pie_chart(portfolio, short_name, threshold)  # 05
    return base64.b64decode(b64) if b64 else None  # 06


def _png_response(png_bytes):  # 07
    bio = BytesIO(png_bytes)  # 08
    resp = send_file(bio, mimetype="image/png")  # 09
    # prevent caching so you always see the latest                   # 10
    resp.headers["Cache-Control"] = "no-store, max-age=0"  # 11
    return resp  # 12


def sort_by_portfolio(hits):
    """Default: sort hits by portfolio name (numeric vs. string)."""

    def portfolio_sort_key(item):
        portfolio_str = str(item["portfolio"]).strip()
        # If it's purely digits, sort numerically; else sort lexically
        if re.match(r"^\d+$", portfolio_str):
            return (0, int(portfolio_str))
        else:
            return (1, portfolio_str)

    return sorted(hits, key=portfolio_sort_key)


def sort_by_class1(hits):
    """Sort hits by the 'percent1' descending."""
    return sorted(hits, key=lambda x: x["percent1"], reverse=True)


def sort_by_class2(hits):
    """Sort hits by the 'percent2' descending."""
    return sorted(hits, key=lambda x: x["percent2"], reverse=True)


def sort_by_sector(hits):
    """Sort hits by 'sectorPercent' descending."""
    return sorted(hits, key=lambda x: x["sectorPercent"], reverse=True)


def sort_by_cash(hits):
    """Sort hits by the total cash in the row's `cashRows` descending."""

    def total_cash(h):
        return sum(r["marketValue"] for r in h["cashRows"])

    return sorted(hits, key=total_cash, reverse=True)


def load_data_if_needed():  # 01
    global data, CSV1_PATH, CSV2_PATH  # 02
    if not (CSV1_PATH and CSV2_PATH):  # 03
        return  # 04
    if data is not None:  # 05
        return  # 06
        # 07
    # --- Read both CSVs (robust encodings) ------------------------------------------- # 08
    df1 = read_csv_any_encoding(CSV1_PATH)  # CSV_ONE (holdings)                       # 09
    df2 = read_csv_any_encoding(CSV2_PATH)  # CSV_TWO (Account # -> Short Name map)    # 10
    # 11
    # --- Resolve columns (tolerant to variants) -------------------------------------- # 12
    acct1_col = find_col(
        df1.columns,
        "Account Number",
        "Account #",
        "Acct #",  # 13
        "Acct Number",
        "AccountNo",
        "Account",
        required=True,
    )  # 14
    acct2_col = find_col(
        df2.columns,
        "Account #",
        "Account Number",
        "Acct #",  # 15
        "Acct Number",
        "AccountNo",
        "Account",
        required=True,
    )  # 16
    short2_col = find_col(
        df2.columns, "Short Name", "ShortName", "Short_Name", required=True  # 17
    )  # 18
    # 19
    # If CSV_ONE already has a short-name-like column, standardize its name ------------# 20
    short1_guess = find_col(df1.columns, "Short Name", "ShortName", "Short_Name")  # 21
    if short1_guess and short1_guess != "Short Name":  # 22
        df1 = df1.rename(columns={short1_guess: "Short Name"})  # 23
        # 24
    # --- Normalize account keys ------------------------------------------------------ # 25
    df1 = df1.copy()
    df2 = df2.copy()  # 26
    df1["__acct_key"] = normalize_account_series(df1[acct1_col])  # 27
    df2["__acct_key"] = normalize_account_series(df2[acct2_col])  # 28
    # 29
    # --- Build clean mapping (first occurrence wins; blank short names treated as NA) -# 30
    map_df = df2.dropna(subset=["__acct_key"])[["__acct_key", short2_col]].drop_duplicates(  # 31
        subset=["__acct_key"], keep="first"  # 32
    )  # 33
    map_df = map_df.rename(columns={short2_col: "Short Name_MAP"})  # 34
    map_df["Short Name_MAP"] = (
        map_df["Short Name_MAP"].astype(str).str.strip().replace({"": pd.NA})  # 35
    )  # 36
    # 37
    # --- Merge and coalesce to a single "Short Name" --------------------------------- # 38
    merged = df1.merge(map_df, on="__acct_key", how="left", suffixes=("", "_DROP"))  # 39
    # if CSV_ONE had a Short Name, it is 'Short Name'; mapping is 'Short Name_MAP'      # 40
    if "Short Name" in merged.columns and "Short Name_MAP" in merged.columns:  # 41
        merged["Short Name"] = merged["Short Name_MAP"].combine_first(merged["Short Name"])  # 42
        merged.drop(columns=["Short Name_MAP"], inplace=True)  # 43
    elif "Short Name_MAP" in merged.columns:  # 44
        merged["Short Name"] = merged["Short Name_MAP"]  # 45
        merged.drop(columns=["Short Name_MAP"], inplace=True)  # 46
    elif "Short Name" not in merged.columns:  # 47
        # neither mapping nor existing column -> create it                              # 48
        merged["Short Name"] = pd.NA  # 49
        # 50
    # Final fallback: if still NA, use the account number string                        # 51
    merged["Short Name"] = merged["Short Name"].fillna(merged[acct1_col].astype(str))  # 52
    # 53
    # Clean temp key                                                                    # 54
    merged.drop(columns=["__acct_key"], inplace=True)  # 55
    # 56
    data = merged  # 57


with app.app_context():
    load_data_if_needed()


@app.route("/", methods=["GET", "POST"])
def index():
    global data, CSV_PATH

    if data is None:
        load_data_if_needed()
        if data is None:
            return "<h1>Client_Data.csv not found next to the executable.</h1>", 503

    # else proceed
    else:
        "<h1>Flask is Running</h1>"

    global col_list, COL_PORTFOLIO, COL_SHORTNAME, COL_CLASS, COL_VALUE, COL_ACCOUNT, CASH_CLASS, COMMON_STOCK, COL_SEGMENT, cached_classes, cached_segments, short_class_names

    col_list = data.columns.tolist()  # 01
    COL_PORTFOLIO = find_col(col_list, "Portfolio Name", "Portfolio", "PortfolioNumber")  # 02
    COL_SHORTNAME = find_col(
        col_list, "Short Name", "ShortName", "Short_Name", required=True  # 03
    )  # 04
    COL_CLASS = find_col(col_list, "Class", "Asset Class", required=True)  # 05
    COL_VALUE = find_col(
        col_list, "Market Value", "MarketValue", "Value", required=True  # 06
    )  # 07
    COL_ACCOUNT = find_col(
        col_list,
        "Account Number",
        "Account #",
        "Acct #",  # 08
        "Acct Number",
        "AccountNo",
        "Account",
        required=True,
    )  # 09
    COL_SEGMENT = find_col(col_list, "Segment", "Sector")  # 10

    cached_classes = data[COL_CLASS].dropna().unique().tolist()

    # Suppose segments that are not in class list is what is wanted
    cached_segments = [
        seg for seg in data[COL_SEGMENT].dropna().unique().tolist() if seg not in cached_classes
    ]

    prime_color_maps(cached_classes, cached_segments)

    short_class_names = [
        "".join(word[0] for word in line.split() if word) for line in cached_classes
    ]

    # Derive label values from the data, but always fall back to safe defaults
    _class_vals = data[COL_CLASS].astype(str).fillna("").tolist()

    # Use what's in the file if present; otherwise keep the defaults defined above
    derived_cash = search_pattern(_class_vals, "CASH AND EQUIVALENTS")
    derived_commonstock = search_pattern(_class_vals, "COMMON STOCK")

    # Make them global so other routes (e.g., /download) see the same values
    global CASH_CLASS, COMMON_STOCK
    CASH_CLASS = derived_cash or CASH_CLASS
    COMMON_STOCK = derived_commonstock or COMMON_STOCK

    total_portfolios = 0
    total_accounts = 0
    message = ""
    hits = []
    image_data = None
    image_data_sector = None
    chart_shortname = ""

    # Track first/second filter
    classFilter1 = ""
    operator1 = "lt"
    targetPercent1 = ""
    classFilter2 = ""
    operator2 = "lt"
    targetPercent2 = ""
    use_second_filter = False

    # Track sector filter
    selected_sector = "(none)"
    operator_sector = "lt"
    targetPercentSector = ""

    last_sort = "portfolio"  # default
    sector = False

    if request.method == "POST":
        action = request.form.get("action", "")

        # Gather filter inputs
        classFilter1 = request.form.get("classFilter1", "").strip()
        operator1 = request.form.get("operator1", "lt").strip()
        targetPercent1 = request.form.get("targetPercent1", "").strip()

        use_second_filter = request.form.get("use_second_filter") == "on"

        if use_second_filter:
            classFilter2 = request.form.get("classFilter2", "").strip()
            operator2 = request.form.get("operator2", "lt").strip()
            targetPercent2 = request.form.get("targetPercent2", "").strip()

        # Gather sector filter

        last_sort = request.form.get("last_sort", "portfolio")

        if action in (
            "filter",
            "sort_class1",
            "sort_class2",
            "sort_sector",
            "sort_cash",
            "view_chart",
            "download_excel",
        ):

            if classFilter1:
                df_filtered = filter_single(data, classFilter1, operator1, targetPercent1)
            else:
                df_filtered = data  # No filtering if no classFilter1 is specified

            if use_second_filter and classFilter2:
                df_filtered = filter_single(df_filtered, classFilter2, operator2, targetPercent2)

            # 2) If either class is COMMON_STOCK and sector is chosen (and has operator/target),
            #    apply the sector filter the same way

            # handle the sector filter

            if classFilter1 == COMMON_STOCK or (use_second_filter and classFilter2 == COMMON_STOCK):
                selected_sector = request.form.get("sectorFilter", "(none)").strip()
                operator_sector = request.form.get("operator_sector", "lt").strip()
                targetPercentSector = request.form.get("targetPercentSector", "").strip()

                if selected_sector != "(none)":
                    df_filtered = filter_by_sector(
                        df_filtered, selected_sector, operator_sector, targetPercentSector
                    )

            # 4) Sector sum (only if selected_sector is not '(none)')
            df_sec = pd.DataFrame()

            if (selected_sector != "(none)") and (selected_sector):
                df_sec = df_filtered[df_filtered[COL_SEGMENT] == selected_sector]
                df_sec_sum = (
                    df_sec.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE]
                    .sum()
                    .rename("sector_value")
                )
                df_sec_sum = df_sec_sum.reset_index()

                sector = True

            # Now build the final hits list with *all* columns:
            # We want to compute percent1, percent2, sectorPercent for each portfolio/shortName
            #   => do a grouped approach to get total, class1 sum, class2 sum, sector sum, etc.
            # Then we can also gather the 'cashRows'.

            if not df_sec.empty:
                total_portfolios = df_sec[COL_PORTFOLIO].nunique()
                total_accounts = df_sec[COL_ACCOUNT].nunique()

            else:
                if not df_filtered.empty:
                    total_portfolios = df_filtered[COL_PORTFOLIO].nunique()
                    total_accounts = df_filtered[COL_ACCOUNT].nunique()

            # 1) total_value by (portfolio, shortName)
            df_port_sum = (
                df_filtered.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE]
                .sum()
                .rename("total_value")
            )
            df_port_sum = df_port_sum.reset_index()

            # 2) class1 sum
            if classFilter1:
                df_class1 = df_filtered[df_filtered[COL_CLASS] == classFilter1]
                df_class1_sum = (
                    df_class1.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE]
                    .sum()
                    .rename("class1_value")
                )
                df_class1_sum = df_class1_sum.reset_index()
            else:
                # empty
                df_class1_sum = pd.DataFrame(columns=[COL_PORTFOLIO, COL_SHORTNAME, "class1_value"])

            # 3) class2 sum
            if use_second_filter and (classFilter2 != "(none)"):
                df_class2 = df_filtered[df_filtered[COL_CLASS] == classFilter2]
                df_class2_sum = (
                    df_class2.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE]
                    .sum()
                    .rename("class2_value")
                )
                df_class2_sum = df_class2_sum.reset_index()
            else:
                df_class2_sum = pd.DataFrame(columns=[COL_PORTFOLIO, COL_SHORTNAME, "class2_value"])

            df_merge = df_port_sum

            if classFilter1:
                df_merge = df_port_sum.merge(
                    df_class1_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how="left"
                )

            if use_second_filter and classFilter2:
                df_merge = df_merge.merge(
                    df_class2_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how="left"
                )

            if sector:
                df_merge = df_merge.merge(df_sec_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how="left")

            df_merge.fillna(0, inplace=True)

            # For each row in df_merge, compute percent1, percent2, sectorPercent
            # Then gather the "cashRows"
            tmp_hits = []
            for idx, row in df_merge.iterrows():
                port = row[COL_PORTFOLIO]
                sname = row[COL_SHORTNAME]
                total_val = row.get("total_value", 0)  # Use .get() to avoid KeyError
                class1_val = row.get("class1_value", 0)
                class2_val = row.get("class2_value", 0)
                sector_val = row.get("sector_value", 0)

                portfoliovalue = total_val  # This should now reflect the merged 'total_value'

                percent1 = (class1_val / total_val) * 100 if total_val else 0
                percent2 = (class2_val / total_val) * 100 if total_val else 0
                sectorPercent = (sector_val / total_val) * 100 if total_val else 0

                # gather cash
                cd = df_filtered[
                    (df_filtered[COL_PORTFOLIO] == port)
                    & (df_filtered[COL_SHORTNAME] == sname)
                    & (df_filtered[COL_CLASS] == CASH_CLASS)
                ]
                c_rows = []
                for _, row_ in cd.iterrows():
                    if row_[COL_VALUE] > 0:
                        c_rows.append(
                            {"accountNumber": row_[COL_ACCOUNT], "marketValue": row_[COL_VALUE]}
                        )

                if use_second_filter:

                    tmp_hits.append(
                        {
                            "portfolio": port,
                            "shortName": sname,
                            "percent1": percent1,
                            "percent2": percent2,
                            "sectorPercent": sectorPercent,
                            "cashRows": c_rows,
                            "portfolioValue": portfoliovalue,  # Correctly reflect total_value here
                        }
                    )

                else:
                    tmp_hits.append(
                        {
                            "portfolio": port,
                            "shortName": sname,
                            "percent1": percent1,
                            "sectorPercent": sectorPercent,
                            "cashRows": c_rows,
                            "portfolioValue": portfoliovalue,  # Correctly reflect total_value here
                        }
                    )

            # Default sort by portfolio
            hits = sort_by_portfolio(tmp_hits)

            if not hits:
                message = "No portfolios met the filter criteria."

        # Then handle sorts and view_chart
        if action == "sort_class1":
            hits = sort_by_class1(hits)
            last_sort = "sort_class1"
        elif action == "sort_class2":
            hits = sort_by_class2(hits)
            last_sort = "sort_class2"
        elif action == "sort_sector":
            hits = sort_by_sector(hits)
            last_sort = "sort_sector"
        elif action == "sort_cash":
            hits = sort_by_cash(hits)
            last_sort = "sort_cash"

        # alter hits
        # Re-apply last sort if user just changed filters (common pattern):
        if last_sort == "sort_class1":
            hits = sort_by_class1(hits)
        elif last_sort == "sort_class2":
            hits = sort_by_class2(hits)
        elif last_sort == "sort_sector":
            hits = sort_by_sector(hits)
        elif last_sort == "sort_cash":
            hits = sort_by_cash(hits)

        if action == "download_excel":
            # 1) Save filter parameters to session, if needed
            session["classFilter1"] = classFilter1
            session["operator1"] = operator1
            session["target_percent1"] = targetPercent1

            session["use_second_filter"] = use_second_filter

            if use_second_filter:
                session["classFilter2"] = classFilter2
                session["operator2"] = operator2
                session["target_percent2"] = targetPercent2

            if classFilter1 == COMMON_STOCK or (use_second_filter and classFilter2 == COMMON_STOCK):
                session["selected_sector"] = selected_sector
                session["operator_sector"] = operator_sector
                session["targetPercentSector"] = targetPercentSector

            # `hits` is probably a list of dicts.
            # Usually JSON-serializable, but let's ensure it is

            # etc.

            # 2) redirect (GET) to the download route
            # need to return template string here?
            return redirect(url_for("download"))

    # --- Row-span helpers for template: recompute on the current order -------------------  # 01
    # Must be placed AFTER all sorting has been applied.                                     # 02
    for h in hits:  # 03
        # rows each (portfolio, shortName) will occupy                                       # 04
        h["short_rowspan"] = max(1, len(h["cashRows"]))  # 05

    # Compute row-spans for each contiguous run of the same portfolio in the current order   # 06
    i = 0  # 07
    n = len(hits)  # 08
    while i < n:  # 09
        port = hits[i]["portfolio"]  # 10
        run_span = 0  # 11
        j = i  # 12
        # advance j while portfolio stays the same (contiguous block)                        # 13
        while j < n and hits[j]["portfolio"] == port:  # 14
            run_span += hits[j]["short_rowspan"]  # 15
            j += 1  # 16

        # mark first row of this contiguous portfolio block                                  # 17
        hits[i]["is_portfolio_first"] = True  # 18
        hits[i]["portfolio_rowspan"] = run_span  # 19

        # mark the rest of this contiguous block                                             # 20
        for k in range(i + 1, j):  # 21
            hits[k]["is_portfolio_first"] = False  # 22
            hits[k]["portfolio_rowspan"] = 0  # 23

        # continue with the next block                                                       # 24
        i = j  # 25
    # -------------------------------------------------------------------------------------  # 26

    csv_one_name = os.path.basename(CSV1_PATH) if CSV1_PATH else CSV_ONE_FILENAME  # 02 NEW
    csv_two_name = os.path.basename(CSV2_PATH) if CSV2_PATH else CSV_TWO_FILENAME  # 03 NEW
    csv_one_created = file_created_str(CSV1_PATH) if CSV1_PATH else "N/A"  # 04 NEW
    csv_two_created = file_created_str(CSV2_PATH) if CSV2_PATH else "N/A"

    # --------------------------------------------------------------------------------------         # 05 NEW

    # Render template
    return render_template_string(
        template,
        cached_classes=cached_classes,
        cached_segments=cached_segments,
        selected_class1=classFilter1,
        operator1=operator1,
        target_percent1=targetPercent1,
        selected_class2=classFilter2,
        operator2=operator2,
        total_portfolios=total_portfolios,
        total_accounts=total_accounts,
        target_percent2=targetPercent2,
        use_second_filter=use_second_filter,
        sector=sector,
        selected_sector=selected_sector,
        operator_sector=operator_sector,
        target_percent_sector=targetPercentSector,
        last_sort=last_sort,
        hits=hits,
        image_data=image_data,
        image_data_sector=image_data_sector,
        chart_shortname=chart_shortname,
        message=message,
        short_class_names=short_class_names,
        COMMON_STOCK=COMMON_STOCK,
        csv_one_name=csv_one_name,  # NEW
        csv_two_name=csv_two_name,  # NEW
        csv_one_created=csv_one_created,  # NEW
        csv_two_created=csv_two_created,  # NEW
    )


@app.get("/chart/class")  # 01
def chart_class():  # 02
    portfolio = request.args.get("portfolio", "").strip()  # 03
    short = request.args.get("short", "").strip()  # 04
    png = _pie_bytes(portfolio, short)  # 05
    if not png:  # 06
        return ("", 204)  # 07
    return _png_response(png)  # 08


@app.get("/chart/sector")  # 09
def chart_sector():  # 10
    portfolio = request.args.get("portfolio", "").strip()  # 11
    short = request.args.get("short", "").strip()  # 12
    png = _sector_pie_bytes(portfolio, short)  # 13
    if not png:  # 14
        return ("", 204)  # 15
    return _png_response(png)  # 16


@app.route("/download")
def download():
    # Ensure data present
    if data is None:
        load_data_if_needed()
        if data is None:
            return "No data file loaded.", 503

    # Pull current filter state from session (the POST that set it happens in /)
    classFilter1 = session.get("classFilter1", "")
    operator1 = session.get("operator1", "lt")
    targetPercent1 = session.get("target_percent1", "")

    use_second_filter = bool(session.get("use_second_filter", False))
    classFilter2 = session.get("classFilter2", "") if use_second_filter else ""
    operator2 = session.get("operator2", "lt")
    targetPercent2 = session.get("target_percent2", "")

    selected_sector = session.get("selected_sector", "(none)")
    operator_sector = session.get("operator_sector", "lt")
    targetPercentSector = session.get("targetPercentSector", "")

    # Recompute filtered dataset
    df_filtered = data
    if classFilter1:
        df_filtered = filter_single(df_filtered, classFilter1, operator1, targetPercent1)
    if use_second_filter and classFilter2:
        df_filtered = filter_single(df_filtered, classFilter2, operator2, targetPercent2)

    # Sector filter (only if COMMON_STOCK is involved and a sector is selected)
    if (classFilter1 == COMMON_STOCK) or (use_second_filter and classFilter2 == COMMON_STOCK):
        if selected_sector and selected_sector != "(none)" and targetPercentSector:
            df_filtered = filter_by_sector(
                df_filtered, selected_sector, operator_sector, targetPercentSector
            )

    # Sector sum if applicable
    sector = False
    df_sec_sum = None
    if selected_sector and selected_sector != "(none)":
        df_sec = df_filtered[df_filtered[COL_SEGMENT] == selected_sector]
        if not df_sec.empty:
            df_sec_sum = (
                df_sec.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE]
                .sum()
                .rename("sector_value")
                .reset_index()
            )
            sector = True

    # Base totals
    df_port_sum = (
        df_filtered.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE]
        .sum()
        .rename("total_value")
        .reset_index()
    )

    # Class sums
    df_merge = df_port_sum.copy()
    if classFilter1:
        df_class1_sum = (
            df_filtered[df_filtered[COL_CLASS] == classFilter1]
            .groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE]
            .sum()
            .rename("class1_value")
            .reset_index()
        )
        df_merge = df_merge.merge(df_class1_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how="left")
    if use_second_filter and classFilter2:
        df_class2_sum = (
            df_filtered[df_filtered[COL_CLASS] == classFilter2]
            .groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE]
            .sum()
            .rename("class2_value")
            .reset_index()
        )
        df_merge = df_merge.merge(df_class2_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how="left")
    if sector and df_sec_sum is not None:
        df_merge = df_merge.merge(df_sec_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how="left")

    df_merge.fillna(0, inplace=True)

    # Build hits like / (so Excel matches the table)
    tmp_hits = []
    for _, row in df_merge.iterrows():
        port = row[COL_PORTFOLIO]
        sname = row[COL_SHORTNAME]
        total_val = row.get("total_value", 0.0)
        class1_val = row.get("class1_value", 0.0)
        class2_val = row.get("class2_value", 0.0)
        sector_val = row.get("sector_value", 0.0)
        percent1 = (class1_val / total_val) * 100 if total_val else 0.0
        percent2 = (class2_val / total_val) * 100 if total_val else 0.0
        sectorPercent = (sector_val / total_val) * 100 if total_val else 0.0

        # cash rows
        cd = df_filtered[
            (df_filtered[COL_PORTFOLIO] == port)
            & (df_filtered[COL_SHORTNAME] == sname)
            & (df_filtered[COL_CLASS] == CASH_CLASS)
        ]
        c_rows = []
        for _, r in cd.iterrows():
            if r[COL_VALUE] > 0:
                c_rows.append({"accountNumber": r[COL_ACCOUNT], "marketValue": r[COL_VALUE]})

        hit = {
            "portfolio": port,
            "shortName": sname,
            "percent1": percent1,
            "percent2": percent2,
            "sectorPercent": sectorPercent,
            "cashRows": c_rows,
            "portfolioValue": total_val,
        }
        tmp_hits.append(hit)

    hits = sort_by_portfolio(tmp_hits)

    # Human-friendly column headers (use the same short tags as the table)
    try:
        class1_header = short_class_names[cached_classes.index(classFilter1)]
    except Exception:
        class1_header = classFilter1 or "Class1"
    class2_header = None
    if use_second_filter and classFilter2:
        try:
            class2_header = short_class_names[cached_classes.index(classFilter2)]
        except Exception:
            class2_header = classFilter2

    # Build rows for Excel
    export_data = []
    for h in hits:
        if h["cashRows"]:
            for cr in h["cashRows"]:
                row = {
                    "Portfolio Name": h["portfolio"],
                    "Short Name": h["shortName"],
                    class1_header: round(h["percent1"], 2),
                    "Portfolio Value": h["portfolioValue"],
                    "Account Number": cr["accountNumber"],
                    "Cash": cr["marketValue"],
                }
                if class2_header:
                    row[class2_header] = round(h["percent2"], 2)
                if sector:
                    row["Sector"] = round(h["sectorPercent"], 2)
                export_data.append(row)
        else:
            row = {
                "Portfolio Name": h["portfolio"],
                "Short Name": h["shortName"],
                class1_header: round(h["percent1"], 2),
                "Portfolio Value": h["portfolioValue"],
                "Account Number": "",
                "Cash": 0,
            }
            if class2_header:
                row[class2_header] = round(h["percent2"], 2)
            if sector:
                row["Sector"] = round(h["sectorPercent"], 2)
            export_data.append(row)

    df_excel = pd.DataFrame(export_data)
    if df_excel.empty:
        # No file to download — tell the browser there is no content
        return ("No data to download for the current filter.", 204)

    # Create Excel file in memory
    output = io.BytesIO()
    df_excel.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    # Format money and percentages
    percent_headers = [class1_header]
    if class2_header:
        percent_headers.append(class2_header)
    if sector:
        percent_headers.append("Sector")

    for col in ws.iter_cols(min_row=1, max_row=1):
        header = col[0].value
        if header in ("Cash", "Portfolio Value"):
            for cell in ws[col[0].column_letter][1:]:
                cell.number_format = "$#,##0.00"
        elif header in percent_headers:
            for cell in ws[col[0].column_letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.value = cell.value / 100.0
                cell.number_format = "0.00%"

    # Widths + center
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[col_letter].width = min(max_len + 5, 60)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="filtered_data.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# add flag --noconsole when building the app with pyinstaller to avoid altogether
@app.route("/favicon.ico")
def favicon():
    return "", 204


def run_flask():
    # Run Flask *without* reloader or debug if you want it to stay quiet
    app.run(debug=False, use_reloader=False)


class MainWindow(QMainWindow):
    def __init__(self, url):
        super().__init__()
        self.setWindowTitle("")

        # 1) Use your custom MyWebEngineView
        self.browser = MyWebEngineView()

        # 3) Set as central widget
        self.setCentralWidget(self.browser)

        # Locate both CSVs next to the executable / script                                # 01
        path1 = self.get_data_file_path(CSV_ONE_FILENAME)  # 02
        path2 = self.get_data_file_path(CSV_TWO_FILENAME)  # 03
        missing = [p for p in (path1, path2) if not os.path.isfile(p)]  # 04

        if missing:  # 05
            self.show_file_error("Missing data file(s):\n" + "\n".join(missing))  # 06
        else:  # 07
            global CSV1_PATH, CSV2_PATH  # 08
            CSV1_PATH, CSV2_PATH = path1, path2  # 09

            # 2) Load the desired URL
            self.browser.load(QUrl(url))

    def check_data_file(self):
        path = self.get_data_file_path()

        if not os.path.isfile(path):

            self.show_file_error(path)
            time.sleep(5)
            sys.exit(1)

    def show_file_error(self, path):
        """
        Display an error dialog if the data file is missing.
        """
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Warning)
        msg.setWindowTitle("Data File Error")
        msg.setText(f"The data file was not found:\n\n{path}")
        msg.setInformativeText("Please place the file in the correct directory and restart.")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()

    def get_data_file_path(self, filename="Client_Data.csv"):
        # If we're running in a PyInstaller bundle
        if getattr(sys, "frozen", False):
            # The executable is here:

            base_path = os.path.dirname(sys.executable)
        else:
            # Running in normal Python
            base_path = os.path.dirname(os.path.abspath(__file__))

        return os.path.join(base_path, filename)


class MyWebEngineView(QWebEngineView):
    def __init__(self, parent=None):
        super().__init__(parent)
        profile = QWebEngineProfile.defaultProfile()
        profile.downloadRequested.connect(self.on_downloadRequested)

    def on_downloadRequested(self, download_item):
        """
        This method is invoked whenever the user tries to download a file
        (i.e. whenever Flask sends a file with Content-Disposition: attachment).

        We must explicitly accept and set the download path.

        """
        print(f"Download requested for: {download_item.url().toString()}")

        # Let user choose path or set it automatically
        filename, _ = QFileDialog.getSaveFileName(
            self, "Save Excel", "filtered_data.xlsx", "Excel Files (*.xlsx);;All Files (*)"
        )
        if filename:
            download_item.setPath(filename)
            download_item.accept()
        else:
            print("did not download file")
            download_item.cancel()


def main():
    flask_thread = threading.Thread(target=run_flask)
    flask_thread.daemon = True
    flask_thread.start()

    app = QApplication(sys.argv)
    main_window = MainWindow("http://127.0.0.1:5000/")
    main_window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
