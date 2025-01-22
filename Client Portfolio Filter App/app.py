import re
import io
import base64
import pandas as pd
from flask import Flask, request, render_template_string, send_file, session, redirect, url_for
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import os
import time

from io import BytesIO
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import json

import sys
import threading
import webbrowser

from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox
from PyQt5.QtWebEngineWidgets import QWebEngineView, QWebEngineProfile
from PyQt5.QtCore import QUrl


app = Flask(__name__)

app.secret_key='secret_key'
# cahnge buttons to reflect input, ensure pie colors are consistent
#portfolios, acc numbers, needing input of optional second filter
# order by class 2 only if use_second_filter is True

#excel not perf
#if neither is common stock, it wont filter (or only 1 filter)
#portfolios, acc numbers
#municipal bonds < 100 doesn't return all accounts
#ensure that both prior pie charts are deleted after each new request
#cookie warning when displaying charts



from flask_cors import CORS
CORS(app)

# def get_data_file_path(filename='Client_Data.csv'):
#     # If we're running in a PyInstaller bundle
#     if getattr(sys, 'frozen', False):
#         # The executable is here:
#         base_path = sys._MEIPASS  # if you had embedded data
#         # But if you want to store the CSV right next to the exe,
#         # you could do:
#         base_path = os.path.dirname(sys.executable)
#     else:
#         # Running in normal Python
#         base_path = os.path.dirname(os.path.abspath(__file__))

#     return os.path.join(base_path, filename)

# csv_path = get_data_file_path()

# # Load your data
# data = pd.read_csv(csv_path)


def search_pattern(strings, pattern):
    """
    Search for a pattern in a list of strings, ignoring non-alphabetical characters
    (except spaces) in the search and making it case-insensitive.

    Parameters:
    - strings: List of strings to search within.
    - pattern: Search pattern containing only letters and spaces.

    Returns:
    - List of strings matching the pattern.
    """
    # Normalize the pattern: remove extra spaces and convert to lowercase
    normalized_pattern = re.sub(r'\s+', ' ', pattern.strip()).lower()
    
    # Prepare a regex to match the normalized pattern
    pattern_regex = re.compile(re.escape(normalized_pattern), re.IGNORECASE)

    # Filter strings for matches
    matches = []
    for string in strings:
        # Normalize the string: keep only letters and spaces, convert to lowercase
        normalized_string = re.sub(r'[^a-zA-Z ]', '', string).lower()
        if pattern_regex.search(normalized_string):
            matches.append(string)
    
    return matches[0]

# --- Columns ---

DATA_FILENAME = "Client_Data.csv"
CSV_PATH = None  # global that gets set once the file is confirmed
data = None      # maybe store a global DataFrame
# --- Build cache lists ---

# ------------- HTML TEMPLATE -------------
from template import template

# ------------- PYTHON HELPER FUNCTIONS -------------
def filter_by_sector(df, sector, operator_, target_str):
    """
    Return only rows in df whose 'Segment' is within the operator_ to target % range
    (just like the other filters). If sector is empty, pass through the df unchanged.
    
    However, it might be more intuitive to FIRST filter out rows that do not match 
    the sector name, and THEN apply a % check. But the user might want 
    lt/gt/eq 30% for the sector. We must compute the actual % for each 
    portfolio and shortName's sector. For simplicity, let's do what we do 
    for the other class filters, but for the chosen sector.

    If user has selected "sector" but no operator or target, pass through anyway.
    """
    if not sector:
        return df  # no sector filter

    try:
        target_percent = float(target_str)
    except ValueError:
        # If user typed something invalid as percent => no results
        return pd.DataFrame()

    # Group at the portfolio level, to find the total
    group_tot = df.groupby(COL_PORTFOLIO)[COL_VALUE].sum().reset_index()
    group_tot_dict = dict(zip(group_tot[COL_PORTFOLIO], group_tot[COL_VALUE]))

    # Group by portfolio & segment
    group_seg = df.groupby([COL_PORTFOLIO, COL_SEGMENT])[COL_VALUE].sum().reset_index()

    # Find those portfolios whose sector portion meets the operator condition
    pass_list = []
    for idx, row in group_seg.iterrows():
        port = row[COL_PORTFOLIO]
        seg_name = row[COL_SEGMENT]
        val = row[COL_VALUE]
        total_val = group_tot_dict.get(port, 0)
        if total_val == 0:
            continue
        actual_percent = (val / total_val) * 100.0
        pass_condition = False
        if seg_name == sector:  # match sector
            if operator_ == 'lt' and actual_percent < target_percent:
                pass_condition = True
            elif operator_ == 'gt' and actual_percent > target_percent:
                pass_condition = True
            elif operator_ == 'eq' and abs(actual_percent - target_percent) <= 3:
                pass_condition = True

        if pass_condition:
            pass_list.append(port)

    pass_set = set(pass_list)
    return df[df[COL_PORTFOLIO].isin(pass_set)]


def filter_single(df, class_filter, operator_, target_str):
    """Apply a single filter (Class, Operator, Target %) to 'df', returning new df or hits."""
    if not class_filter:
        return df  # No filter => pass-through
    try:
        target_percent = float(target_str)
    except ValueError:
        # If user typed something invalid as percent => no results
        return pd.DataFrame()

    # Group to compute each portfolio's total
    group_pc = df.groupby([COL_PORTFOLIO, COL_SHORTNAME, COL_CLASS])[COL_VALUE].sum().reset_index()
    total_by_portfolio = df.groupby(COL_PORTFOLIO)[COL_VALUE].sum().to_dict()

    pass_list = []
    for idx, row in group_pc.iterrows():
        if row[COL_CLASS] == class_filter:
            port = row[COL_PORTFOLIO]
            val = row[COL_VALUE]
            tot = total_by_portfolio.get(port, 0)
            if tot == 0:
                continue
            actual_percent = (val / tot) * 100.0

            pass_condition = False
            if operator_ == 'lt':
                pass_condition = (actual_percent < target_percent)
            elif operator_ == 'gt':
                pass_condition = (actual_percent > target_percent)
            elif operator_ == 'eq':
                # ±3% tolerance
                pass_condition = (abs(actual_percent - target_percent) <= 3)

            if pass_condition:
                pass_list.append(port)

    pass_set = set(pass_list)
    return df[df[COL_PORTFOLIO].isin(pass_set)] #add sum here and concatenate somehow?


def apply_both_filters(class1, op1, pct1, class2, op2, pct2, use_second, df):
    """Apply up to 2 filters in an AND fashion on df. If use_second=True, apply second also."""
    df1 = filter_single(df, class1, op1, pct1)

    if not use_second:

        total_by_portfolio = df1.groupby(COL_PORTFOLIO)[COL_VALUE].sum().rename('Portfolio Value')
        df1 = df1.merge(total_by_portfolio, on=COL_PORTFOLIO, how='left')

        return df1
    # else, filter the result again
    df2 = filter_single(df1, class2, op2, pct2)

    # Calculate portfolio value
    total_by_portfolio = df2.groupby(COL_PORTFOLIO)[COL_VALUE].sum().rename('Portfolio Value')
    df2 = df2.merge(total_by_portfolio, on=COL_PORTFOLIO, how='left')

    return df2

def convert_to_percent(data_list):
    """Convert a list of values to %, rounded to 2 decimal places."""
    total = sum(data_list)
    return [round((val / total) * 100.0, 2) for val in data_list]


#anti-aliasing figure renderer
matplotlib.use('Agg')

def generate_pie_chart(portfolio, shortName=None, threshold=3):
    """Generate a pie chart of Class breakdown for a portfolio (and optional shortName)."""

    class_color_map = {}
    for i, cls in enumerate(cached_classes):
        class_color_map[cls] = list(mcolors.TABLEAU_COLORS.values())[i % len(mcolors.TABLEAU_COLORS)]

    colormap = class_color_map

    df_filtered = data[data[COL_PORTFOLIO] == portfolio]
    if shortName:
        df_filtered = df_filtered[df_filtered[COL_SHORTNAME] == shortName]

    if df_filtered.empty:
        return None

    class_values = df_filtered.groupby(COL_CLASS)[COL_VALUE].sum()
    fig, ax = plt.subplots(figsize=(4,4))

    colors = [colormap[cls] for cls in class_values.index]

    wedges, _texts, autotexts = ax.pie(
        class_values.values,
        labels=None,
        autopct=lambda p: f'{p:.1f}%' if p >= threshold else '',
        startangle=90,
        colors=colors
    )
    ax.axis('equal')
    #ax.set_title(f"Asset Breakdown - {shortName}")

    fig.subplots_adjust(bottom=0.3) 

    legend_labels = [
    f"{label} ({size}%)" if size <= threshold else f"{label}"
    for label, size in zip(class_values.index, convert_to_percent(class_values.values))
    ]

    ax.legend(
        wedges,
        legend_labels,
        title="Class",
        loc='upper center',
        bbox_to_anchor=(0.5, -0.1),
        fancybox=True,
        shadow=True,
        ncol=2
    )

    buf = io.BytesIO()
    fig.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode('utf-8')


def generate_sector_pie_chart(portfolio, short_name=None, threshold=2.5):
    """Among COMMON STOCK holdings, break down by Segment (Sector) in a separate pie."""
    
    segment_color_map = {}
    for i, seg in enumerate(cached_segments):
        segment_color_map[seg] = list(mcolors.TABLEAU_COLORS.values())[(i+10) % len(mcolors.TABLEAU_COLORS)]
    
    colormap=segment_color_map

    df_filtered = data[
        (data[COL_PORTFOLIO] == portfolio) &
        (data[COL_CLASS] == COMMON_STOCK)
    ]
    if short_name:
        df_filtered = df_filtered[df_filtered[COL_SHORTNAME] == short_name]

    if df_filtered.empty:
        return None

    sector_values = df_filtered.groupby(COL_SEGMENT)[COL_VALUE].sum()
    fig, ax = plt.subplots(figsize=(4, 4))

    colors = [colormap[cls] for cls in sector_values.index]

    wedges, _txt, autotxt = ax.pie(
        sector_values.values,
        labels=None,
        autopct=lambda p: f'{p:.1f}%' if p >= threshold else '',
        startangle=90,
        colors=colors
    )
    ax.axis('equal')
    #ax.set_title(f"Common Stock by Sector - {short_name}")
    fig.subplots_adjust(bottom=0.3)
    legend_labels = [
    f"{label} ({size}%)" if size <= threshold else f"{label}"
    for label, size in zip(sector_values.index, convert_to_percent(sector_values.values))
    ]
    ax.legend(
        wedges,
        legend_labels,
        title="Segment",
        loc='upper center',
        bbox_to_anchor=(0.5, -0.1),
        ncol=2
    )

    buf = io.BytesIO()
    fig.savefig(buf, format='png', bbox_inches='tight')
    buf.seek(0)
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode('utf-8')


def sort_by_portfolio(hits):
    """Default: sort hits by portfolio name (numeric vs. string)."""
    def portfolio_sort_key(item):
        portfolio_str = str(item['portfolio']).strip()
        # If it's purely digits, sort numerically; else sort lexically
        if re.match(r'^\d+$', portfolio_str):
            return (0, int(portfolio_str))
        else:
            return (1, portfolio_str)
    return sorted(hits, key=portfolio_sort_key)


def sort_by_class1(hits):
    """Sort hits by the 'percent1' descending."""
    return sorted(hits, key=lambda x: x['percent1'], reverse=True)

def sort_by_class2(hits):
    """Sort hits by the 'percent2' descending."""
    return sorted(hits, key=lambda x: x['percent2'], reverse=True)

def sort_by_sector(hits):
    """Sort hits by 'sectorPercent' descending."""
    return sorted(hits, key=lambda x: x['sectorPercent'], reverse=True)

def sort_by_cash(hits):
    """Sort hits by the total cash in the row's `cashRows` descending."""
    def total_cash(h):
        return sum(r['marketValue'] for r in h['cashRows'])
    return sorted(hits, key=total_cash, reverse=True)


@app.before_first_request
def load_data_if_needed():
    global data, CSV_PATH
    if CSV_PATH and data is None:
        data = pd.read_csv(CSV_PATH)

@app.route('/', methods=['GET', 'POST'])
def index():

    if data is None:
        "<h1>Data is missing or wasn't loaded!</h1>"
        time.sleep(5)
        sys.exit(1)
        
# else proceed
    else:
        "<h1>Flask is Running</h1>"

    
    global col_list, COL_PORTFOLIO, COL_SHORTNAME, COL_CLASS, COL_VALUE, COL_ACCOUNT, CASH_CLASS, COMMON_STOCK, COL_SEGMENT, cached_classes, cached_segments, short_class_names

    col_list = data.columns.tolist()
    COL_PORTFOLIO = search_pattern(col_list, 'Portfolio Name')
    COL_SHORTNAME = search_pattern(col_list, 'Short Name')
    COL_CLASS = search_pattern(col_list, 'Class')
    COL_VALUE = search_pattern(col_list, 'Market Value')
    COL_ACCOUNT = search_pattern(col_list, 'Account Number')
    CASH_CLASS = search_pattern(data[COL_CLASS].tolist(), 'CASH AND EQUIVALENTS')
    COMMON_STOCK = search_pattern(data[COL_CLASS].tolist(), 'COMMON STOCK')
    COL_SEGMENT = search_pattern(col_list, 'Segment') 

    cached_classes = data[COL_CLASS].dropna().unique().tolist()

    # Suppose you only want segments that are NOT in your class list
    cached_segments = [
        seg for seg in data[COL_SEGMENT].dropna().unique().tolist()
        if seg not in cached_classes
    ]

    short_class_names = [
        ''.join(word[0] for word in line.split() if word)
        for line in cached_classes
    ]
    


    total_portfolios = 0
    total_accounts = 0
    message = ""
    hits = []
    image_data = None
    image_data_sector = None
    chart_shortname = ""

    # Track first/second filter
    classFilter1 = ""
    operator1 = "lt"
    targetPercent1 = ""
    classFilter2 = ""
    operator2 = "lt"
    targetPercent2 = ""
    use_second_filter = False

    # Track sector filter
    selected_sector = '(none)'
    operator_sector = "lt"
    targetPercentSector = ""

    last_sort = "portfolio"  # default
    sector = False

    if request.method == 'POST':
        action = request.form.get('action', '')

        # Initialize session storage for charts
        if 'chart_data' not in session:
            session['chart_data'] = {}
            session['chart_data_sector'] = {}

        # Gather filter inputs
        classFilter1 = request.form.get('classFilter1', '').strip()
        operator1 = request.form.get('operator1', 'lt').strip()
        targetPercent1 = request.form.get('targetPercent1', '').strip()

        use_second_filter = request.form.get('use_second_filter') == 'on'

        if use_second_filter:
            classFilter2 = request.form.get('classFilter2', '').strip()
            operator2 = request.form.get('operator2', 'lt').strip()
            targetPercent2 = request.form.get('targetPercent2', '').strip()


    
        # Gather sector filter

        last_sort = request.form.get('last_sort', 'portfolio')


        if action in ('filter', 'sort_class1', 'sort_class2', 'sort_sector', 'sort_cash', 'view_chart', 'download_excel'):
            
            if classFilter1:
                df_filtered = filter_single(data, classFilter1, operator1, targetPercent1)
            else:
                df_filtered = data  # No filtering if no classFilter1 is specified

            if use_second_filter and classFilter2:
                df_filtered = filter_single(df_filtered, classFilter2, operator2, targetPercent2)

            # 2) If either class is COMMON_STOCK and sector is chosen (and has operator/target),
            #    apply the sector filter the same way

            #handle the sector filter 
           
            if (classFilter1 == COMMON_STOCK or (use_second_filter and classFilter2 == COMMON_STOCK)):
                selected_sector = request.form.get('sectorFilter', '(none)').strip()
                operator_sector = request.form.get('operator_sector', 'lt').strip()
                targetPercentSector = request.form.get('targetPercentSector', '').strip()   

                if selected_sector != '(none)':
                    df_filtered = filter_by_sector(df_filtered, selected_sector, operator_sector, targetPercentSector)
            
            # 4) Sector sum (only if selected_sector is not '(none)')
            df_sec = pd.DataFrame()
            

            if (selected_sector != '(none)') and (selected_sector):
                df_sec = df_filtered[df_filtered[COL_SEGMENT] == selected_sector]
                df_sec_sum = df_sec.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE].sum().rename('sector_value')
                df_sec_sum = df_sec_sum.reset_index()

                sector = True
                
            # Now build the final hits list with *all* columns:
            # We want to compute percent1, percent2, sectorPercent for each portfolio/shortName
            #   => do a grouped approach to get total, class1 sum, class2 sum, sector sum, etc.
            # Then we can also gather the 'cashRows'.
        
            
            if not df_sec.empty:
                total_portfolios = df_sec[COL_PORTFOLIO].nunique()
                total_accounts = df_sec[COL_ACCOUNT].nunique()

            else:
                if not df_filtered.empty:
                    total_portfolios = df_filtered[COL_PORTFOLIO].nunique()
                    total_accounts = df_filtered[COL_ACCOUNT].nunique()

            # 1) total_value by (portfolio, shortName)
            df_port_sum = df_filtered.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE].sum().rename('total_value')
            df_port_sum = df_port_sum.reset_index()

            # 2) class1 sum
            if classFilter1:
                df_class1 = df_filtered[df_filtered[COL_CLASS] == classFilter1]
                df_class1_sum = df_class1.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE].sum().rename('class1_value')
                df_class1_sum = df_class1_sum.reset_index()
            else:
                # empty
                df_class1_sum = pd.DataFrame(columns=[COL_PORTFOLIO, COL_SHORTNAME, 'class1_value'])

            # 3) class2 sum
            if use_second_filter and (classFilter2 != '(none)'):
                df_class2 = df_filtered[df_filtered[COL_CLASS] == classFilter2]
                df_class2_sum = df_class2.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE].sum().rename('class2_value')
                df_class2_sum = df_class2_sum.reset_index()
            else:
                df_class2_sum = pd.DataFrame(columns=[COL_PORTFOLIO, COL_SHORTNAME, 'class2_value'])


            df_merge=df_port_sum
            
            if classFilter1:
                df_merge = df_port_sum.merge(df_class1_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how='left')

            if use_second_filter and classFilter2:
                df_merge = df_merge.merge(df_class2_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how='left')
            
            if sector:
                df_merge = df_merge.merge(df_sec_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how='left')

            df_merge.fillna(0, inplace=True)

            # For each row in df_merge, compute percent1, percent2, sectorPercent
            # Then gather the "cashRows"
            tmp_hits = []
            for idx, row in df_merge.iterrows():
                port = row[COL_PORTFOLIO]
                sname = row[COL_SHORTNAME]
                total_val = row.get('total_value', 0)  # Use .get() to avoid KeyError
                class1_val = row.get('class1_value', 0)
                class2_val = row.get('class2_value', 0)
                sector_val = row.get('sector_value', 0)

                portfoliovalue = total_val  # This should now reflect the merged 'total_value'

                percent1 = (class1_val / total_val) * 100 if total_val else 0
                percent2 = (class2_val / total_val) * 100 if total_val else 0
                sectorPercent = (sector_val / total_val) * 100 if total_val else 0

                # gather cash
                cd = df_filtered[
                    (df_filtered[COL_PORTFOLIO] == port) &
                    (df_filtered[COL_SHORTNAME] == sname) &
                    (df_filtered[COL_CLASS] == CASH_CLASS)
                ]
                c_rows = []
                for _, row_ in cd.iterrows():
                    if row_[COL_VALUE] > 0:
                        c_rows.append({
                            'accountNumber': row_[COL_ACCOUNT],
                            'marketValue': row_[COL_VALUE]
                        })
                
                if use_second_filter:

                    tmp_hits.append({
                        'portfolio': port,
                        'shortName': sname,
                        'percent1': percent1,
                        'percent2': percent2,
                        'sectorPercent': sectorPercent,
                        'cashRows': c_rows,
                        'portfolioValue': portfoliovalue  # Correctly reflect total_value here
                    })

                else:
                    tmp_hits.append({
                        'portfolio': port,
                        'shortName': sname,
                        'percent1': percent1,
                        'sectorPercent': sectorPercent,
                        'cashRows': c_rows,
                        'portfolioValue': portfoliovalue  # Correctly reflect total_value here
                    })

            # Default sort by portfolio
            hits = sort_by_portfolio(tmp_hits)

            if not hits:
                message = "No portfolios met the filter criteria."

            
        # Then handle sorts and view_chart
        if action == 'sort_class1':
            hits = sort_by_class1(hits)
            last_sort = 'sort_class1'
        elif action == 'sort_class2':
            hits = sort_by_class2(hits)
            last_sort = 'sort_class2'
        elif action == 'sort_sector':
            hits = sort_by_sector(hits)
            last_sort = 'sort_sector'
        elif action == 'sort_cash':
            hits = sort_by_cash(hits)
            last_sort = 'sort_cash'
        
        # Pie chart generation
        if action == 'view_chart' or action in ('filter', 'sort_class1', 'sort_class2', 'sort_sector', 'sort_cash'):
            chart_portfolio = request.form.get('chart_portfolio', '').strip()
            chart_shortname = request.form.get('chart_shortname', '').strip()

            # Check if chart data already exists in the session
            if chart_portfolio and chart_shortname:
                if action == 'view_chart' or not session['chart_data']:
                    image_data = generate_pie_chart(chart_portfolio, chart_shortname)
                    image_data_sector = generate_sector_pie_chart(chart_portfolio, chart_shortname)

                    # Save generated charts to session
                    session['chart_data'] = image_data
                    session['chart_data_sector'] = image_data_sector

            # Retrieve the chart data from the session
            image_data = session.get('chart_data', None)
            image_data_sector = session.get('chart_data_sector', None)

            if not image_data and not image_data_sector:
                message = f"No data found for portfolio={chart_portfolio}"
            elif not image_data:
                message = f"No asset breakdown data found for portfolio={chart_portfolio}"
            elif not image_data_sector:
                message = f"No sector breakdown data found for portfolio={chart_portfolio}"
        

        #alter hits
        # Re-apply last sort if user just changed filters (common pattern):
        if last_sort == 'sort_class1':
            hits = sort_by_class1(hits)
        elif last_sort == 'sort_class2':
            hits = sort_by_class2(hits)
        elif last_sort == 'sort_sector':
            hits = sort_by_sector(hits)
        elif last_sort == 'sort_cash':
            hits = sort_by_cash(hits)
    
        if action == 'download_excel':
            # 1) Save filter parameters to session, if needed
            session['classFilter1'] = classFilter1
            session['operator1'] = operator1
            session['target_percent1'] = targetPercent1

            session["use_second_filter"] = use_second_filter

            if use_second_filter:
                session['classFilter2'] = classFilter2
                session['operator2'] = operator2
                session['target_percent2'] = targetPercent2

            if (classFilter1 == COMMON_STOCK or (use_second_filter and classFilter2 == COMMON_STOCK)):
                session['selected_sector'] = selected_sector
                session['operator_sector'] = operator_sector
                session['targetPercentSector'] = targetPercentSector
            
            # `hits` is probably a list of dicts. 
            # Usually JSON-serializable, but let's ensure it is
            
            # etc.

            # 2) redirect (GET) to the download route
            #need to return template string here?
            return redirect(url_for('download'))

    
    # Render template
    return render_template_string(
        template,
        cached_classes=cached_classes,
        cached_segments=cached_segments,
        selected_class1=classFilter1,
        operator1=operator1,
        target_percent1=targetPercent1,
        selected_class2=classFilter2,
        operator2=operator2,
        total_portfolios=total_portfolios,
        total_accounts=total_accounts,
        target_percent2=targetPercent2,
        use_second_filter=use_second_filter,\
        sector=sector,
        selected_sector=selected_sector,
        operator_sector=operator_sector,
        target_percent_sector=targetPercentSector,
        last_sort=last_sort,
        hits=hits,
        image_data=image_data,
        image_data_sector=image_data_sector,
        chart_shortname=chart_shortname,
        message=message,
        short_class_names=short_class_names
    )


@app.route('/download')
def download(classFilter1="", operator1="lt", targetPercent1="", use_second_filter=False, classFilter2="", operator2="lt", targetPercent2="", selected_sector='(none)', operator_sector="lt", targetPercentSector="", sector=False ):
    
    export_data = []

    classFilter1 = session['classFilter1']
    operator1 = session['operator1'] 
    targetPercent1 = session['target_percent1'] 

    use_second_filter = session["use_second_filter"] 

    if use_second_filter:
        classFilter2 = session['classFilter2']
        operator2 = session['operator2'] 
        targetPercent2 = session['target_percent2']

    if (classFilter1 == COMMON_STOCK or (use_second_filter and classFilter2 == COMMON_STOCK)):
        selected_sector = session['selected_sector']
        operator_sector = session['operator_sector'] 
        targetPercentSector = session['targetPercentSector']


    if classFilter1:
        df_filtered = filter_single(data, classFilter1, operator1, targetPercent1)
    else:
        df_filtered = data  # No filtering if no classFilter1 is specified

    if use_second_filter and classFilter2:
        df_filtered = filter_single(df_filtered, classFilter2, operator2, targetPercent2)

    # 2) If either class is COMMON_STOCK and sector is chosen (and has operator/target),
    #    apply the sector filter the same way

    #handle the sector filter 

    if (classFilter1 == COMMON_STOCK or (use_second_filter and classFilter2 == COMMON_STOCK)):
        selected_sector = request.form.get('sectorFilter', '(none)').strip()
        operator_sector = request.form.get('operator_sector', 'lt').strip()
        targetPercentSector = request.form.get('targetPercentSector', '').strip()

        if selected_sector != '(none)':
            df_filtered = filter_by_sector(df_filtered, selected_sector, operator_sector, targetPercentSector)
    
    # 4) Sector sum (only if selected_sector is not '(none)')
    df_sec = pd.DataFrame()
    

    if (selected_sector != '(none)') and (selected_sector):
        df_sec = df_filtered[df_filtered[COL_SEGMENT] == selected_sector]
        df_sec_sum = df_sec.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE].sum().rename('sector_value')
        df_sec_sum = df_sec_sum.reset_index()

        sector = True
        
    # Now build the final hits list with *all* columns:
    # We want to compute percent1, percent2, sectorPercent for each portfolio/shortName
    #   => do a grouped approach to get total, class1 sum, class2 sum, sector sum, etc.
    # Then we can also gather the 'cashRows'.


    # 1) total_value by (portfolio, shortName)
    df_port_sum = df_filtered.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE].sum().rename('total_value')
    df_port_sum = df_port_sum.reset_index()

    # 2) class1 sum
    if classFilter1:
        df_class1 = df_filtered[df_filtered[COL_CLASS] == classFilter1]
        df_class1_sum = df_class1.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE].sum().rename('class1_value')
        df_class1_sum = df_class1_sum.reset_index()
    else:
        # empty
        df_class1_sum = pd.DataFrame(columns=[COL_PORTFOLIO, COL_SHORTNAME, 'class1_value'])

    # 3) class2 sum
    if use_second_filter and (classFilter2 != '(none)'):
        df_class2 = df_filtered[df_filtered[COL_CLASS] == classFilter2]
        df_class2_sum = df_class2.groupby([COL_PORTFOLIO, COL_SHORTNAME])[COL_VALUE].sum().rename('class2_value')
        df_class2_sum = df_class2_sum.reset_index()
    else:
        df_class2_sum = pd.DataFrame(columns=[COL_PORTFOLIO, COL_SHORTNAME, 'class2_value'])


    df_merge=df_port_sum
    
    if classFilter1:
        df_merge = df_port_sum.merge(df_class1_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how='left')

    if use_second_filter and classFilter2:
        df_merge = df_merge.merge(df_class2_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how='left')
    
    if sector:
        df_merge = df_merge.merge(df_sec_sum, on=[COL_PORTFOLIO, COL_SHORTNAME], how='left')

    df_merge.fillna(0, inplace=True)

    # For each row in df_merge, compute percent1, percent2, sectorPercent
    # Then gather the "cashRows"
    tmp_hits = []
    for idx, row in df_merge.iterrows():
        port = row[COL_PORTFOLIO]
        sname = row[COL_SHORTNAME]
        total_val = row.get('total_value', 0)  # Use .get() to avoid KeyError
        class1_val = row.get('class1_value', 0)
        class2_val = row.get('class2_value', 0)
        sector_val = row.get('sector_value', 0)

        portfoliovalue = total_val  # This should now reflect the merged 'total_value'

        percent1 = (class1_val / total_val) * 100 if total_val else 0
        percent2 = (class2_val / total_val) * 100 if total_val else 0
        sectorPercent = (sector_val / total_val) * 100 if total_val else 0

        # gather cash
        cd = df_filtered[
            (df_filtered[COL_PORTFOLIO] == port) &
            (df_filtered[COL_SHORTNAME] == sname) &
            (df_filtered[COL_CLASS] == CASH_CLASS)
        ]
        c_rows = []
        for _, row_ in cd.iterrows():
            if row_[COL_VALUE] > 0:
                c_rows.append({
                    'accountNumber': row_[COL_ACCOUNT],
                    'marketValue': row_[COL_VALUE]
                })
        
        if use_second_filter:

            tmp_hits.append({
                'portfolio': port,
                'shortName': sname,
                'percent1': percent1,
                'percent2': percent2,
                'sectorPercent': sectorPercent,
                'cashRows': c_rows,
                'portfolioValue': portfoliovalue  # Correctly reflect total_value here
            })

        else:
            tmp_hits.append({
                'portfolio': port,
                'shortName': sname,
                'percent1': percent1,
                'sectorPercent': sectorPercent,
                'cashRows': c_rows,
                'portfolioValue': portfoliovalue  # Correctly reflect total_value here
            })

    # Default sort by portfolio
    hits = sort_by_portfolio(tmp_hits)

    class1_name=cached_classes.index(classFilter1)

    try:
        class2_name = cached_classes.index(classFilter2)
    except Exception:
        class2_name = '(none)'
        
    for h in hits:
        # If no cashRows, we still want one row (with blank or None for account/cash)
        if h['cashRows']:
            for cr in h['cashRows']:
                if use_second_filter & sector:
                    export_data.append({
                        'Portfolio Name': h['portfolio'],
                        'Short Name': h['shortName'],
                        class1_name: round(h['percent1'],2),
                        class2_name: round(h['percent2'],2),
                        'Sector': round(h['sectorPercent'],2),
                        'Account Number': cr['accountNumber'],
                        'Cash': cr['marketValue'],
                        'Portfolio Value': h['portfolioValue']
                    })
                elif use_second_filter:

                    export_data.append({
                        'Portfolio Name': h['portfolio'],
                        'Short Name': h['shortName'],
                        class1_name: round(h['percent1'],2),
                        class2_name: round(h['percent2'],2),
                        'Account Number': cr['accountNumber'],
                        'Cash': cr['marketValue'],
                        'Portfolio Value': h['portfolioValue']
                    })
                elif sector:

                    export_data.append({
                        'Portfolio Name': h['portfolio'],
                        'Short Name': h['shortName'],
                        class1_name: round(h['percent1'],2),
                        'Sector': round(h['sectorPercent'],2),
                        'Account Number': cr['accountNumber'],
                        'Cash': cr['marketValue'],
                        'Portfolio Value': h['portfolioValue']
                    })

                else:
                    export_data.append({
                        'Portfolio Name': h['portfolio'],
                        'Short Name': h['shortName'],
                        class1_name: round(h['percent1'],2),
                        'Account Number': cr['accountNumber'],
                        'Cash': cr['marketValue'],
                        'Portfolio Value': h['portfolioValue']
                    })

        else:
            # No cashRows => one row with None
            if use_second_filter & sector:
                export_data.append({
                    'Portfolio Name': h['portfolio'],
                    'Short Name': h['shortName'],
                    class1_name: round(h['percent1'],2),
                    class2_name: round(h['percent2'],2),
                    'Sector': round(h['sectorPercent'],2),
                    'Account Number': cr['accountNumber'],
                    'Cash': cr['marketValue'],
                    'Portfolio Value': h['portfolioValue']
                })
            elif use_second_filter:

                export_data.append({
                    'Portfolio Name': h['portfolio'],
                    'Short Name': h['shortName'],
                    class1_name: round(h['percent1'],2),
                    class2_name: round(h['percent2'],2),
                    'Account Number': cr['accountNumber'],
                    'Cash': cr['marketValue'],
                    'Portfolio Value': h['portfolioValue']
                })
            elif sector:

                export_data.append({
                    'Portfolio Name': h['portfolio'],
                    'Short Name': h['shortName'],
                    class1_name: round(h['percent1'],2),
                    'Sector': round(h['sectorPercent'],2),
                    'Account Number': cr['accountNumber'],
                    'Cash': cr['marketValue'],
                    'Portfolio Value': h['portfolioValue']
                })

            else:
                export_data.append({
                    'Portfolio Name': h['portfolio'],
                    'Short Name': h['shortName'],
                    class1_name: round(h['percent1'],2),
                    'Account Number': cr['accountNumber'],
                    'Cash': cr['marketValue'],
                    'Portfolio Value': h['portfolioValue']
                })

    df_excel = pd.DataFrame(export_data)
    output = BytesIO()


    if df_excel.empty:
        # If there's nothing to download
        message = "No data to download."

    else:
        # Generate an Excel file in memory
        output = io.BytesIO()
        df_excel.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)
        
        #format column width:
        
        # Load the workbook and select the active sheet
        wb = load_workbook(output)
        ws = wb.active

        for col in ws.iter_cols(min_row=1, max_row=1):  # Iterate through the first row (headers)
            header = col[0].value  # Get the column header
            if header in ["Cash", "Portfolio Value"]:  # Assuming "Cash" column
                for cell in ws[col[0].column_letter]:  # Iterate through all cells in this column
                    cell.number_format = "$#,##0.00"
            elif header in [class1_name, class2_name, "Sector"]:  # Percentage columns
                for cell in ws[col[0].column_letter][1:]:  # Iterate through all cells in this column
                    if cell.value is not None:
                        cell.value = cell.value / 100  # Convert to decimal
                    cell.number_format = "0.00%"

        # Adjust column widths and apply centering
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Get the column letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                # Apply center alignment
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col[0].value in [class1_name, class2_name, "Sector"]:  # Percentage columns
                ws.column_dimensions[col_letter].width = max_length - 8  # something weird happens with percent formatted columns
            else:
                ws.column_dimensions[col_letter].width = max_length + 5 # Add some padding

        # Save the workbook back to BytesIO
        output = BytesIO()  # Reset the BytesIO object
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='filtered_data.xlsx',  # or 'whatever_you_like.xlsx'
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

@app.route('/favicon.ico')
def favicon():
    return "", 204


def run_flask():
    # Run Flask *without* reloader or debug if you want it to stay quiet
    app.run(debug=False, use_reloader=False)

class MainWindow(QMainWindow):
    def __init__(self, url):
        super().__init__()
        self.setWindowTitle("")

        # 1) Use your custom MyWebEngineView
        self.browser = MyWebEngineView()

        # 2) Load the desired URL
        self.browser.load(QUrl(url))

        # 3) Set as central widget
        self.setCentralWidget(self.browser)

        path = self.get_data_file_path()
        if not os.path.isfile(path):
            self.show_file_error(path)
            # Possibly exit or let the user close
        else:
            # 2) Store path in a global so Flask can load it
            global CSV_PATH
            CSV_PATH = path

    def check_data_file(self):
        path = self.get_data_file_path()
    
        if not os.path.isfile(path):

            self.show_file_error(path)
            time.sleep(5)
            sys.exit(1)

    def show_file_error(self, path):
        """
        Display an error dialog if the data file is missing.
        """
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Warning)
        msg.setWindowTitle("Data File Error")
        msg.setText(f"The data file was not found:\n\n{path}")
        msg.setInformativeText("Please place the file in the correct directory and restart.")
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()

    def get_data_file_path(self, filename='Client_Data.csv'):
        # If we're running in a PyInstaller bundle
        if getattr(sys, 'frozen', False):
            # The executable is here:

            base_path = os.path.dirname(sys.executable)
        else:
            # Running in normal Python
            base_path = os.path.dirname(os.path.abspath(__file__))

        return os.path.join(base_path, filename)

class MyWebEngineView(QWebEngineView):
    def __init__(self, parent=None):
        super().__init__(parent)
        profile = QWebEngineProfile.defaultProfile()
        profile.downloadRequested.connect(self.on_downloadRequested)

    def on_downloadRequested(self, download_item):
        """
        This method is invoked whenever the user tries to download a file
        (i.e. whenever Flask sends a file with Content-Disposition: attachment).

        We must explicitly accept and set the download path.

        """
        print(f"Download requested for: {download_item.url().toString()}")
        
        # Let user choose path or set it automatically
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel",
            "filtered_data.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        if filename:
            download_item.setPath(filename)
            download_item.accept()
        else:
            print('did not download file')
            download_item.cancel()

def main():
    flask_thread = threading.Thread(target=run_flask)
    flask_thread.daemon = True
    flask_thread.start()

    app = QApplication(sys.argv)
    main_window = MainWindow("http://127.0.0.1:5000/")
    main_window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
































